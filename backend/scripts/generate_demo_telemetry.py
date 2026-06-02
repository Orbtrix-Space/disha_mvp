"""
DISHA · MONITOR — demo telemetry generator.

Produces `netra_demo_telemetry.xlsx` with three sheets:
  telemetry          — 360 samples × 23 channels, 10 s cadence
  packet_definitions — TM packet schema (field name, type, length)
  scenario_log       — narrative reference (operator-facing)

The signal is hand-shaped to walk the dashboard through a scripted
operational story:

    0–8   nominal cruise
    8–15  eclipse #1 + comms drift (SNR 18→6)
   15–22  comms DEGRADED
   22–28  recovery + payload ACTIVE downlink
   28–35  eclipse #2 + thermal warning
   35–42  battery anomaly (faster discharge)
   42–48  combined NO_CONTACT + low SOC (critical)
   48–55  autonomy intervention (payload safed)
   55–60  recovery to nominal

Every continuous channel carries small Gaussian noise so traces look
like real telemetry, not perfect curves. Discrete states transition
over ~30 s via smoothstep shoulders rather than stepping instantly.

Output paths:
  samples/netra_demo_telemetry.xlsx                (canonical source)
  frontend/public/samples/netra_demo_telemetry.xlsx (served by Vite)
"""

from __future__ import annotations

import math
import os
import random
import shutil
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─── Config ──────────────────────────────────────────────────────────

T0 = datetime(2026, 6, 2, 8, 0, 0, tzinfo=timezone.utc)
DT_SEC = 10
N_SAMPLES = 360                      # 60 min @ 10 s cadence
CAPACITY_WH = 500.0
SEED = 42

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir))
OUT_SRC  = os.path.join(REPO_ROOT, "samples", "netra_demo_telemetry.xlsx")
OUT_PUB  = os.path.join(REPO_ROOT, "frontend", "public", "samples",
                        "netra_demo_telemetry.xlsx")


# ─── Smoothing helpers ───────────────────────────────────────────────

def smoothstep(x: float, a: float, b: float) -> float:
    if x <= a:
        return 0.0
    if x >= b:
        return 1.0
    t = (x - a) / (b - a)
    return t * t * (3 - 2 * t)


def pulse(x: float, lo: float, hi: float, shoulder: float = 0.5) -> float:
    """Smooth pulse: 1 inside [lo, hi], soft shoulders of width `shoulder` (min)."""
    return smoothstep(x, lo - shoulder, lo + shoulder) \
         - smoothstep(x, hi - shoulder, hi + shoulder)


def low_pass(prev, target, alpha=0.05):
    return prev + alpha * (target - prev)


# ─── Channel definitions (used by telemetry + packet_definitions) ────

CHANNELS = [
    ("timestamp_utc",          "string",  20),
    ("mission_elapsed_s",      "uint16",   2),
    ("battery_soc_pct",        "float",    4),
    ("battery_voltage_v",      "float",    4),
    ("battery_current_a",      "float",    4),
    ("solar_power_w",          "float",    4),
    ("power_margin_w",         "float",    4),
    ("temp_battery_c",         "float",    4),
    ("temp_payload_c",         "float",    4),
    ("temp_panel_c",           "float",    4),
    ("comms_snr_db",           "float",    4),
    ("comms_link_state",       "enum",     1),
    ("comms_ber",              "float",    4),
    ("adcs_mode",              "enum",     1),
    ("adcs_pointing_err_deg",  "float",    4),
    ("adcs_omega_dps",         "float",    4),
    ("payload_state",          "enum",     1),
    ("payload_power_w",        "float",    4),
    ("eps_bus_voltage_v",      "float",    4),
    ("eps_load_w",             "float",    4),
    ("storage_used_pct",       "float",    4),
    ("eclipse_flag",           "uint8",    1),
    ("ground_station",         "string",  20),
]


SCENARIO = [
    (" 0–8",  "Nominal cruise",
     "All subsystems green, risk ~10%, mode NOMINAL"),
    (" 8–15", "Eclipse #1 + comms drift",
     "Solar→0, battery discharging, SNR drops 18→6 dB"),
    ("15–22", "Comms degradation event",
     "Link DEGRADED, SNR <5 dB, Comms-1 fires, risk → 35%"),
    ("22–28", "Comms recovery + downlink active",
     "Link NOMINAL, payload ACTIVE, storage fills"),
    ("28–35", "Eclipse #2 + thermal warning",
     "Panel temp swings, battery temp 22→34°C, Thermal-2 fires"),
    ("35–42", "Battery anomaly (faster discharge)",
     "SOC drops faster than profile, Power-3 fires, risk → 55%"),
    ("42–48", "Critical: NO_CONTACT + low SOC",
     "Multiple rules fire, autonomy → GUARDED, risk → 75%"),
    ("48–55", "Autonomy intervention",
     "Payload SAFE, load drops, recovery begins"),
    ("55–60", "Recovery to nominal",
     "Mode → NOMINAL, risk back to ~15%"),
]


# ─── Signal generator ────────────────────────────────────────────────

def generate_rows():
    rng = random.Random(SEED)
    rows = []

    # Stateful values (low-pass / integrators)
    soc = 92.0
    storage = 35.0
    batt_temp = 22.0
    panel_temp = 25.0
    payload_temp = 12.0

    for i in range(N_SAMPLES):
        t_sec = i * DT_SEC
        t_min = t_sec / 60.0

        # ── Eclipse: pulses at 8-15 and 28-35 with 30s shoulders ────
        eclipse = pulse(t_min, 8.0, 15.0, 0.5) + pulse(t_min, 28.0, 35.0, 0.5)
        eclipse_flag = 1 if eclipse > 0.5 else 0
        sun = max(0.0, 1.0 - eclipse)

        # ── Solar power ───────────────────────────────────────────────
        solar = 170.0 * sun + rng.gauss(0.0, 1.8)
        solar = max(0.0, solar)

        # ── Payload state / power ─────────────────────────────────────
        if 22.0 <= t_min < 35.0:
            payload_active = True
        else:
            payload_active = False
        if 48.0 <= t_min < 55.0:
            payload_active = False   # safed by autonomy intervention

        payload_state = "ACTIVE" if payload_active else "IDLE"
        payload_power = max(0.0, 12.0 + rng.gauss(0.0, 0.4)) if payload_active else 0.0

        # ── EPS load ──────────────────────────────────────────────────
        eps_load = 18.0 + payload_power + rng.gauss(0.0, 0.35)
        # Anomaly window 35–42 — a stuck heater driving extra load
        if 35.0 <= t_min < 42.0:
            eps_load += 3.5 * smoothstep(t_min, 35.0, 36.5)

        # ── Power margin + SOC integration ────────────────────────────
        power_margin = solar - eps_load
        d_wh = (solar - eps_load) * (DT_SEC / 3600.0)
        soc += (d_wh / CAPACITY_WH) * 100.0
        # Extra discharge during the anomaly window
        if 35.0 <= t_min < 42.0:
            soc -= 0.045
        # Critical: 42–48 NO_CONTACT, battery already low — slight extra drain
        if 42.0 <= t_min < 48.0:
            soc -= 0.025
        soc = max(0.0, min(100.0, soc))

        # ── Battery voltage tracks SOC ────────────────────────────────
        batt_v = 11.5 + (soc / 100.0) * 1.1 + rng.gauss(0.0, 0.015)
        batt_v = max(11.0, min(12.7, batt_v))

        # ── Battery current ───────────────────────────────────────────
        batt_i = (solar - eps_load) / 12.0 + rng.gauss(0.0, 0.08)
        batt_i = max(-3.0, min(5.0, batt_i))

        # ── Thermal — low-pass with target ────────────────────────────
        # Battery temp target: 22°C nominal, rises 22→34 during 28–35
        batt_target = 22.0
        if 28.0 <= t_min < 35.0:
            batt_target = 22.0 + 12.0 * smoothstep(t_min, 28.0, 34.0)
        elif 35.0 <= t_min < 50.0:
            # slow cool-down
            batt_target = 34.0 - 14.0 * smoothstep(t_min, 38.0, 50.0)
        batt_temp = low_pass(batt_temp, batt_target, alpha=0.06)
        batt_temp += rng.gauss(0.0, 0.12)

        # Panel temp: 50°C in sun, -20°C in eclipse, big swing with lag
        panel_target = -20.0 + 70.0 * sun
        panel_temp = low_pass(panel_temp, panel_target, alpha=0.08)
        panel_temp += rng.gauss(0.0, 0.4)

        # Payload temp: warms during ACTIVE
        payload_target = 12.0 + (28.0 if payload_active else 0.0)
        payload_temp = low_pass(payload_temp, payload_target, alpha=0.05)
        payload_temp += rng.gauss(0.0, 0.25)

        # ── Comms / link state ────────────────────────────────────────
        if t_min < 8.0:
            snr_target, link, gs = 16.0, "NOMINAL", "ISTRAC_Bangalore"
        elif 8.0 <= t_min < 15.0:
            # smooth drift 18 → 6
            snr_target = 18.0 - 12.0 * smoothstep(t_min, 8.0, 15.0)
            link = "DEGRADED" if snr_target < 8.0 else "NOMINAL"
            gs = "ISTRAC_Bangalore"
        elif 15.0 <= t_min < 22.0:
            snr_target = 4.0 + 1.6 * math.sin((t_min - 15.0) * 2.4)
            link = "DEGRADED"
            gs = "ISTRAC_Bangalore"
        elif 22.0 <= t_min < 28.0:
            snr_target = 6.0 + 10.0 * smoothstep(t_min, 22.0, 24.0)
            link = "NOMINAL" if snr_target > 8.0 else "DEGRADED"
            gs = "Port_Blair"
        elif 28.0 <= t_min < 35.0:
            snr_target = 15.0 + rng.gauss(0.0, 0.5)
            link = "NOMINAL"
            gs = "Port_Blair"
        elif 35.0 <= t_min < 42.0:
            snr_target = 13.0 + 1.0 * math.sin(t_min)
            link = "NOMINAL"
            gs = "Port_Blair"
        elif 42.0 <= t_min < 48.0:
            snr_target = 0.0
            link = "NO_CONTACT"
            gs = ""
        elif 48.0 <= t_min < 55.0:
            snr_target = 2.0 + 10.0 * smoothstep(t_min, 48.0, 50.0)
            link = "DEGRADED" if snr_target < 8.0 else "NOMINAL"
            gs = "ISTRAC_Bangalore"
        else:
            snr_target = 15.0
            link = "NOMINAL"
            gs = "ISTRAC_Bangalore"

        snr = max(0.0, snr_target + rng.gauss(0.0, 0.25))

        # BER: log-spaced — better link, lower BER
        if link == "NOMINAL":
            ber_exp = -8.0 + rng.gauss(0.0, 0.3)
        elif link == "DEGRADED":
            ber_exp = -4.5 + rng.gauss(0.0, 0.4)
        else:
            ber_exp = -3.0
        ber = 10.0 ** ber_exp

        # ── ADCS ──────────────────────────────────────────────────────
        if 42.0 <= t_min < 55.0:
            adcs_mode = "SAFE"
        elif 28.0 <= t_min < 35.0 and eclipse_flag:
            adcs_mode = "SUN_POINT"
        else:
            adcs_mode = "NADIR"

        # Pointing baseline 0.3°; spikes at mode transitions
        pointing = 0.3 + abs(rng.gauss(0.0, 0.08))
        for tm in (8.0, 15.0, 22.0, 28.0, 35.0, 42.0, 48.0, 55.0):
            d = abs(t_min - tm)
            if d < 0.5:
                pointing += 6.0 * (1.0 - d / 0.5)
        pointing = max(0.0, min(10.0, pointing))

        # Angular rate
        omega = 0.05 + abs(rng.gauss(0.0, 0.015))
        if pointing > 1.0:
            omega += 0.4 + abs(rng.gauss(0.0, 0.1))
        omega = min(2.0, omega)

        # ── Storage ───────────────────────────────────────────────────
        if payload_active:
            storage += 0.18 + rng.gauss(0.0, 0.02)
        storage = max(30.0, min(95.0, storage))

        # ── EPS bus voltage ───────────────────────────────────────────
        eps_v = batt_v - 0.1 + rng.gauss(0.0, 0.04)
        eps_v = max(11.0, min(13.0, eps_v))

        rows.append({
            "timestamp_utc":         (T0 + timedelta(seconds=t_sec))
                                      .strftime("%Y-%m-%dT%H:%M:%SZ"),
            "mission_elapsed_s":     t_sec,
            "battery_soc_pct":       round(soc, 2),
            "battery_voltage_v":     round(batt_v, 2),
            "battery_current_a":     round(batt_i, 2),
            "solar_power_w":         round(solar, 1),
            "power_margin_w":        round(power_margin, 1),
            "temp_battery_c":        round(batt_temp, 1),
            "temp_payload_c":        round(payload_temp, 1),
            "temp_panel_c":          round(panel_temp, 1),
            "comms_snr_db":          round(snr, 2),
            "comms_link_state":      link,
            "comms_ber":             float(f"{ber:.2e}"),
            "adcs_mode":             adcs_mode,
            "adcs_pointing_err_deg": round(pointing, 2),
            "adcs_omega_dps":        round(omega, 3),
            "payload_state":         payload_state,
            "payload_power_w":       round(payload_power, 1),
            "eps_bus_voltage_v":     round(eps_v, 2),
            "eps_load_w":            round(eps_load, 1),
            "storage_used_pct":      round(storage, 1),
            "eclipse_flag":          eclipse_flag,
            "ground_station":        gs,
        })
    return rows


# ─── Workbook builder ────────────────────────────────────────────────

THIN = Side(style="thin", color="222226")
BORDER_BOTTOM = Border(bottom=THIN)


def write_workbook(rows, path):
    wb = Workbook()

    # Sheet 1 — telemetry
    ws = wb.active
    ws.title = "telemetry"
    headers = [c[0] for c in CHANNELS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="C8C8CC", name="Consolas", size=10)
        cell.fill = PatternFill("solid", fgColor="111114")
        cell.alignment = Alignment(horizontal="left")
        cell.border = BORDER_BOTTOM
    for row in rows:
        ws.append([row[h] for h in headers])
    # Width hints
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = max(13, len(h) + 2)

    # Sheet 2 — packet_definitions
    ws2 = wb.create_sheet("packet_definitions")
    ws2.append(["field_name", "type", "length_bytes"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="C8C8CC", name="Consolas", size=10)
        cell.fill = PatternFill("solid", fgColor="111114")
        cell.border = BORDER_BOTTOM
    for name, ftype, length in CHANNELS:
        ws2.append([name, ftype, length])
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 14

    # Sheet 3 — scenario_log
    ws3 = wb.create_sheet("scenario_log")
    ws3.append(["time_min", "event", "expected_reaction"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="C8C8CC", name="Consolas", size=10)
        cell.fill = PatternFill("solid", fgColor="111114")
        cell.border = BORDER_BOTTOM
    for t, ev, react in SCENARIO:
        ws3.append([t, ev, react])
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 38
    ws3.column_dimensions["C"].width = 70

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


# ─── CLI ─────────────────────────────────────────────────────────────

def main():
    rows = generate_rows()
    write_workbook(rows, OUT_SRC)
    os.makedirs(os.path.dirname(OUT_PUB), exist_ok=True)
    shutil.copyfile(OUT_SRC, OUT_PUB)

    # Quick sanity print
    last = rows[-1]
    print(f"wrote {len(rows)} rows × {len(CHANNELS)} channels")
    print(f"  → {OUT_SRC}")
    print(f"  → {OUT_PUB}")
    print("final sample:", {
        "soc": last["battery_soc_pct"],
        "v":   last["battery_voltage_v"],
        "snr": last["comms_snr_db"],
        "link": last["comms_link_state"],
        "mode": last["adcs_mode"],
    })


if __name__ == "__main__":
    main()
